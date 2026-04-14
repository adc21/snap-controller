"""
app/ui/export_dialog.py
解析結果エクスポートダイアログ。

完了済みケースの結果を CSV または Excel (.xlsx) 形式で出力します。

レイアウト:
  ┌──────────────────────────────────────────┐
  │ [出力形式: CSV / Excel]                  │
  │ [ケース選択チェックリスト]               │
  │ [出力先パス] [参照…]                     │
  │ [エクスポート] [閉じる]                  │
  └──────────────────────────────────────────┘

UX改善② 第5回 (export_dialog.py):
  ライブ選択サマリー行 + エクスポート完了後「フォルダを開く」ボタン追加。
  チェックボックスの変化に連動して「X ケース選択中（約Y行のデータ）」を
  リアルタイムに表示します。0件選択時はエクスポートボタンを自動無効化します。
  エクスポート成功後の完了ダイアログに「📁 フォルダを開く」ボタンを追加し、
  出力先フォルダをワンクリックでエクスプローラー（Windows）/ Finder（macOS）で開けます。
  また、既存ファイルへの上書きを事前に検出して確認ダイアログを表示します。
"""

from __future__ import annotations

import csv
import logging
import os
import subprocess
import sys
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QVBoxLayout,
    QWidget,
)

from app.models import AnalysisCase, AnalysisCaseStatus

# 応答値の定義 (key, 日本語ラベル, 単位)
_RESPONSE_ITEMS = [
    ("max_disp",        "最大応答相対変位",    "m"),
    ("max_vel",         "最大応答相対速度",    "m/s"),
    ("max_acc",         "最大応答絶対加速度",  "m/s²"),
    ("max_story_disp",  "最大層間変形",        "m"),
    ("max_story_drift", "最大層間変形角",      "rad"),
    ("shear_coeff",     "せん断力係数",        "—"),
    ("max_otm",         "最大転倒モーメント",  "kN·m"),
]


class ExportDialog(QDialog):
    """
    解析結果エクスポートダイアログ。

    Parameters
    ----------
    cases : list of AnalysisCase
        プロジェクト内の全ケースリスト（完了済みのみ表示）。
    default_dir : str
        デフォルトの出力先ディレクトリ。
    parent : QWidget, optional
    """

    def __init__(
        self,
        cases: List[AnalysisCase],
        default_dir: str = "",
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("解析結果のエクスポート")
        self.setMinimumWidth(520)

        self._completed_cases: List[AnalysisCase] = [
            c for c in cases
            if c.status == AnalysisCaseStatus.COMPLETED and c.result_summary
        ]
        self._default_dir = default_dir or str(Path.home())
        self._checkboxes: List[tuple[QCheckBox, AnalysisCase]] = []

        self._setup_ui()
        self._rebuild_checklist()

    # ------------------------------------------------------------------
    # UI Construction
    # ------------------------------------------------------------------

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        # ---- 出力形式 ----
        fmt_row = QHBoxLayout()
        fmt_row.addWidget(QLabel("出力形式:"))
        self._fmt_combo = QComboBox()
        self._fmt_combo.addItems(["CSV (.csv)", "Excel (.xlsx)"])
        self._fmt_combo.currentIndexChanged.connect(self._update_output_path_ext)
        fmt_row.addWidget(self._fmt_combo)
        fmt_row.addStretch()
        layout.addLayout(fmt_row)

        # ---- ケース選択 ----
        group = QGroupBox("エクスポートするケース")
        group_layout = QVBoxLayout(group)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(180)
        self._check_container = QWidget()
        self._check_layout = QVBoxLayout(self._check_container)
        self._check_layout.setAlignment(Qt.AlignTop)
        self._check_layout.setSpacing(2)
        scroll.setWidget(self._check_container)
        group_layout.addWidget(scroll)

        sel_row = QHBoxLayout()
        btn_all = QPushButton("全選択")
        btn_all.setMaximumWidth(70)
        btn_all.clicked.connect(self._select_all)
        btn_none = QPushButton("全解除")
        btn_none.setMaximumWidth(70)
        btn_none.clicked.connect(self._deselect_all)
        sel_row.addWidget(btn_all)
        sel_row.addWidget(btn_none)
        sel_row.addStretch()
        group_layout.addLayout(sel_row)
        layout.addWidget(group)

        # ---- 出力先パス ----
        path_row = QHBoxLayout()
        path_row.addWidget(QLabel("出力先:"))
        self._path_edit = QLineEdit()
        self._path_edit.setPlaceholderText("出力ファイルパスを指定してください")
        default_name = "snap_results.csv"
        self._path_edit.setText(str(Path(self._default_dir) / default_name))
        browse_btn = QPushButton("参照…")
        browse_btn.setMaximumWidth(64)
        browse_btn.clicked.connect(self._browse_output)
        path_row.addWidget(self._path_edit)
        path_row.addWidget(browse_btn)
        layout.addLayout(path_row)

        # ---- UX改善② 第5回: ライブ選択サマリー行 ----
        self._summary_label = QLabel("0 ケース選択中")
        self._summary_label.setStyleSheet(
            "QLabel { color: #555; font-size: 11px; padding: 2px 4px; }"
        )
        layout.addWidget(self._summary_label)

        # ---- ボタン ----
        btn_box = QDialogButtonBox()
        self._export_btn = QPushButton("エクスポート")
        self._export_btn.setDefault(True)
        btn_box.addButton(self._export_btn, QDialogButtonBox.AcceptRole)
        btn_box.addButton(QDialogButtonBox.Close)
        btn_box.accepted.connect(self._do_export)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    # ------------------------------------------------------------------
    # Checklist
    # ------------------------------------------------------------------

    def _rebuild_checklist(self) -> None:
        """完了済みケースのチェックボックスリストを構築します。"""
        for cb, _ in self._checkboxes:
            cb.deleteLater()
        self._checkboxes.clear()

        if not self._completed_cases:
            lbl = QLabel("<i>エクスポート可能なケースがありません（解析完了ケースなし）</i>")
            self._check_layout.addWidget(lbl)
            self._export_btn.setEnabled(False)
            self._update_summary_label()
            return

        for case in self._completed_cases:
            cb = QCheckBox(case.name)
            cb.setChecked(True)
            # UX改善② 第5回: チェック変化でサマリーラベルを更新
            cb.toggled.connect(self._update_summary_label)
            self._check_layout.addWidget(cb)
            self._checkboxes.append((cb, case))

        self._update_summary_label()

    # UX改善② 第5回: ライブ選択サマリー更新
    def _update_summary_label(self) -> None:
        """選択中のケース数と推定データ行数をサマリーラベルに反映します。"""
        selected = [case for cb, case in self._checkboxes if cb.isChecked()]
        count = len(selected)
        if count == 0:
            self._summary_label.setText("⚠ ケースが選択されていません")
            self._summary_label.setStyleSheet(
                "QLabel { color: #c62828; font-size: 11px; padding: 2px 4px; }"
            )
            self._export_btn.setEnabled(False)
        else:
            # 各ケースのresult_dataから大まかな行数を推定（フロア数 × 応答値数）
            total_floors = 0
            for case in selected:
                if case.result_summary:
                    rd = case.result_summary.get("result_data", {})
                    if rd:
                        total_floors += max(len(v) for v in rd.values()) if rd else 0
            row_hint = f"・約 {total_floors} 行" if total_floors > 0 else ""
            self._summary_label.setText(
                f"✓ {count} ケース選択中{row_hint}（エクスポート対象）"
            )
            self._summary_label.setStyleSheet(
                "QLabel { color: #2e7d32; font-size: 11px; padding: 2px 4px; }"
            )
            self._export_btn.setEnabled(True)

    def _select_all(self) -> None:
        for cb, _ in self._checkboxes:
            cb.setChecked(True)
        self._update_summary_label()

    def _deselect_all(self) -> None:
        for cb, _ in self._checkboxes:
            cb.setChecked(False)
        self._update_summary_label()

    # ------------------------------------------------------------------
    # Path helpers
    # ------------------------------------------------------------------

    def _update_output_path_ext(self) -> None:
        """出力形式に応じてファイル拡張子を更新します。"""
        current = self._path_edit.text().strip()
        p = Path(current) if current else Path(self._default_dir) / "snap_results"
        ext = ".xlsx" if self._fmt_combo.currentIndex() == 1 else ".csv"
        self._path_edit.setText(str(p.with_suffix(ext)))

    def _browse_output(self) -> None:
        is_excel = self._fmt_combo.currentIndex() == 1
        if is_excel:
            path, _ = QFileDialog.getSaveFileName(
                self, "エクスポート先を選択",
                self._path_edit.text(),
                "Excel ファイル (*.xlsx);;すべてのファイル (*)"
            )
        else:
            path, _ = QFileDialog.getSaveFileName(
                self, "エクスポート先を選択",
                self._path_edit.text(),
                "CSV ファイル (*.csv);;すべてのファイル (*)"
            )
        if path:
            self._path_edit.setText(path)

    # ------------------------------------------------------------------
    # Export logic
    # ------------------------------------------------------------------

    def _do_export(self) -> None:
        selected = [case for cb, case in self._checkboxes if cb.isChecked()]
        if not selected:
            QMessageBox.warning(self, "警告", "エクスポートするケースを選択してください。")
            return

        output_path = self._path_edit.text().strip()
        if not output_path:
            QMessageBox.warning(self, "警告", "出力先ファイルパスを指定してください。")
            return

        # UX改善② 第5回: 上書き確認
        out_p = Path(output_path)
        if out_p.exists():
            reply = QMessageBox.question(
                self, "上書き確認",
                f"指定のファイルはすでに存在します:\n{output_path}\n\n上書きしますか？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return

        is_excel = self._fmt_combo.currentIndex() == 1

        try:
            if is_excel:
                self._export_excel(selected, output_path)
            else:
                self._export_csv(selected, output_path)

            # UX改善② 第5回: 完了後「フォルダを開く」ボタン付きメッセージボックス
            out_dir = str(Path(output_path).parent)
            msg = QMessageBox(self)
            msg.setWindowTitle("エクスポート完了")
            msg.setText(f"エクスポートが完了しました:\n{output_path}")
            msg.setIcon(QMessageBox.Information)
            open_folder_btn = msg.addButton("📁 フォルダを開く", QMessageBox.ActionRole)
            msg.addButton(QMessageBox.Ok)
            msg.exec()
            if msg.clickedButton() == open_folder_btn:
                self._open_folder(out_dir)
        except ImportError:
            QMessageBox.critical(
                self, "エラー",
                "Excel エクスポートには openpyxl が必要です。\n"
                "pip install openpyxl でインストールしてください。"
            )
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"エクスポートに失敗しました:\n{e}")

    @staticmethod
    def _open_folder(folder_path: str) -> None:
        """
        UX改善② 第5回: 出力先フォルダをOSのファイルマネージャーで開きます。
        Windows: explorer, macOS: open, Linux: xdg-open
        """
        try:
            if sys.platform == "win32":
                os.startfile(folder_path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder_path])
            else:
                subprocess.Popen(["xdg-open", folder_path])
        except Exception:
            logger.debug("フォルダを開けませんでした: %s", folder_path)

    @staticmethod
    def _export_csv(cases: List[AnalysisCase], path: str) -> None:
        """
        ケース結果を CSV 形式でエクスポートします。

        出力形式:
          - サマリーシート相当: ケース名, 各最大値...
          - 層別シート相当: ケース名, 層番号, 各応答値...
        """
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)

        # --- サマリー CSV ---
        summary_path = out.with_stem(out.stem + "_summary")
        with open(summary_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            header = ["ケース名"] + [
                f"{label}[{unit}]" for _, label, unit in _RESPONSE_ITEMS
            ]
            writer.writerow(header)
            for case in cases:
                row = [case.name]
                for key, _, _ in _RESPONSE_ITEMS:
                    row.append(case.result_summary.get(key, ""))
                writer.writerow(row)

        # --- 層別 CSV ---
        floors_path = out.with_stem(out.stem + "_floors")
        with open(floors_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            header = ["ケース名", "層番号"] + [
                f"{label}[{unit}]" for _, label, unit in _RESPONSE_ITEMS
            ]
            writer.writerow(header)
            for case in cases:
                result_data = case.result_summary.get("result_data", {})
                # 全応答値に共通する層番号を収集
                all_floors: set = set()
                for key, _, _ in _RESPONSE_ITEMS:
                    all_floors.update(result_data.get(key, {}).keys())
                for floor in sorted(all_floors):
                    row = [case.name, floor]
                    for key, _, _ in _RESPONSE_ITEMS:
                        row.append(result_data.get(key, {}).get(floor, ""))
                    writer.writerow(row)

    @staticmethod
    def _export_excel(cases: List[AnalysisCase], path: str) -> None:
        """
        ケース結果を Excel (.xlsx) 形式でエクスポートします。

        シート構成:
          - "サマリー": ケース毎の最大値一覧
          - "層別応答値": 全ケース・全層の詳細データ
        """
        import openpyxl  # noqa: import-outside-toplevel
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa

        wb = openpyxl.Workbook()

        # ---- サマリーシート ----
        ws_sum = wb.active
        ws_sum.title = "サマリー"

        header = ["ケース名"] + [
            f"{label}\n[{unit}]" for _, label, unit in _RESPONSE_ITEMS
        ]
        ws_sum.append(header)
        # ヘッダー書式
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_sum[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws_sum.row_dimensions[1].height = 36

        for case in cases:
            row = [case.name]
            for key, _, _ in _RESPONSE_ITEMS:
                val = case.result_summary.get(key, None)
                row.append(round(val, 6) if isinstance(val, float) else val)
            ws_sum.append(row)

        # 列幅調整
        ws_sum.column_dimensions["A"].width = 24
        for col in "BCDEFGH":
            ws_sum.column_dimensions[col].width = 18

        # ---- 層別応答値シート ----
        ws_fl = wb.create_sheet("層別応答値")
        fl_header = ["ケース名", "層番号"] + [
            f"{label}\n[{unit}]" for _, label, unit in _RESPONSE_ITEMS
        ]
        ws_fl.append(fl_header)
        for cell in ws_fl[1]:
            cell.fill = PatternFill("solid", fgColor="70AD47")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_fl.row_dimensions[1].height = 36

        for case in cases:
            result_data = case.result_summary.get("result_data", {})
            all_floors: set = set()
            for key, _, _ in _RESPONSE_ITEMS:
                all_floors.update(result_data.get(key, {}).keys())
            for floor in sorted(all_floors):
                row = [case.name, floor]
                for key, _, _ in _RESPONSE_ITEMS:
                    val = result_data.get(key, {}).get(floor, None)
                    row.append(round(val, 6) if isinstance(val, float) else val)
                ws_fl.append(row)

        ws_fl.column_dimensions["A"].width = 24
        ws_fl.column_dimensions["B"].width = 8
        for col in "CDEFGHI":
            ws_fl.column_dimensions[col].width = 18

        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
